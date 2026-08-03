# -*- coding: utf-8 -*-
"""Генератор лендинга kompanpavel.com/rgb из Excel-таблицы.

Собирает два варианта:
  rgb/        — тёмный (в стилистике kompanpavel.com)
  rgb-light/  — светлый editorial (по структуре пресс-релиза)
"""
import openpyxl, html, json, re, os

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # корень репо
SRC = os.path.join(BASE, "Книги и журналы -2.xlsx")   # исходная таблица (не в репо)
VARIANTS = [
    ("template.html", "rgb"),
    ("template_light.html", "rgb-light"),
    ("template_light.html", "docs"),   # GitHub Pages — копия светлой версии
]

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["Книги"]

def clean_rec(s):
    s = (s or "").strip()
    s = s.replace("фотонАлексей", "фотограф Алексей")  # опечатка в таблице
    s = re.sub(r"\s+", " ", s)
    return s

ROLES = ["журналист", "продюсер", "фотограф", "визажист", "стилист", "основатель Lebigmag"]

def split_rec(s):
    for role in ROLES:
        if s.lower().startswith(role.lower()):
            return role, s[len(role):].strip()
    return "", s

# В таблице сначала идут книги, затем блоком журналы. У книг описание, как
# правило, начинается с отдельной строки-заголовка — но признак неточен с обеих
# сторон: у дописанных в начало «500 страниц ответов» и «Before Now» такой
# строки нет (а это книги), и наоборот, у пары журналов она есть.
# Надёжнее опереться на структуру: находим сплошной участок строк-заголовков —
# он и есть блок книг. Всё до него (дописанное сверху) — тоже книги,
# всё после — журналы.
def _has_head(r):
    return "\n" in str(ws.cell(r, 2).value or "")

_rows = [r for r in range(2, ws.max_row + 1)
         if str(ws.cell(r, 1).value or "").strip()]
_first_head = next((r for r in _rows if _has_head(r)), None)
if _first_head is None:
    LAST_BOOK_ROW = 1
else:
    LAST_BOOK_ROW = _first_head
    for r in _rows:
        if r > _first_head and _has_head(r) and r == LAST_BOOK_ROW + 1:
            LAST_BOOK_ROW = r

# Название + автор в одном заголовке — как у остальных изданий (там имя
# фотографа зашито в названии альбома). У этих двух книг в таблице имени
# автора в названии нет, дописываем вручную по просьбе заказчика:
# сначала название, потом автор (не как обычно — не наоборот). Ключ — idx.
TITLE_OVERRIDES = {
    1: "500 страниц ответов. Том 1 Компан Павел",
    2: "Before Now Данил Головкин",
}

items = []
# диапазон берём по факту: заказчик дописывает строки, зашивать число нельзя
for r in range(2, ws.max_row + 1):
    name = str(ws.cell(r, 1).value or "").strip()
    if not name:
        continue
    desc = str(ws.cell(r, 2).value or "").strip()
    rec = clean_rec(str(ws.cell(r, 4).value or ""))
    kind = "book" if r <= LAST_BOOK_ROW else "mag"
    if "\n" in desc:
        first, rest = desc.split("\n", 1)
        title = first.strip().rstrip(".")
        text = re.sub(r"\s*\n\s*", " ", rest).strip()
    else:
        title = name
        text = desc
    role, person = split_rec(rec)
    idx = r - 1
    title = TITLE_OVERRIDES.get(idx, title)
    # ПЛЕЙСХОЛДЕРЫ шифров — заменить на реальные из РГБ
    code = f"ЕБШ{10876 + idx}"
    items.append(dict(idx=idx, title=title, text=text, code=code,
                      kind=kind, role=role, person=person))

experts = sorted({(i["person"], i["role"]) for i in items})

def card(i):
    kind_label = "Книга" if i["kind"] == "book" else "Журнал"
    t = html.escape(i["title"])
    return f'''
      <article class="card" data-kind="{i['kind']}" data-person="{html.escape(i['person'])}"
               data-search="{html.escape(i['title'].lower())}" id="item-{i['idx']}">
        <div class="card-cover">
          <img src="assets/covers/{i['idx']:02d}.jpg" alt="{t} — обложка" loading="lazy">
          <span class="card-kind">{kind_label}</span>
        </div>
        <div class="card-body">
          <h3 class="card-title">{t}</h3>
          <p class="card-rec">Рекомендует <b>{html.escape(i['person'])}</b><span class="card-role">{html.escape(i['role'])}</span></p>
          <p class="card-text">{html.escape(i['text'])}</p>
          <button class="card-more" type="button" aria-expanded="false">Читать полностью</button>
        </div>
      </article>'''

cards_html = "\n".join(card(i) for i in items)

# лента обложек в hero: 14 изданий, дубль для бесшовной прокрутки.
# 1 — «500 страниц ответов», книга самого KOMPAN PAVEL, по просьбе клиента идёт
# первой. 2 — «Before Now» Головкина: в каталоге есть, но на первый экран его
# клиент просил не выводить. Остальные номера сдвинуты на +2 после того, как
# в таблицу добавили две книги в начало.
MARQUEE_PICKS = [1, 3, 4, 5, 6, 7, 12, 14, 16, 22, 24, 26, 30, 37]
by_idx = {i["idx"]: i for i in items}
mq_one = "\n".join(
    f'      <a class="mq-item" href="#item-{n}" title="{html.escape(by_idx[n]["title"])} — открыть в каталоге">'
    f'<img src="assets/covers/{n:02d}.jpg" alt="{html.escape(by_idx[n]["title"])}" decoding="async"></a>'
    for n in MARQUEE_PICKS if n in by_idx
)
marquee_html = mq_one + "\n" + mq_one

persons_options = "\n".join(
    f'          <option value="{html.escape(p)}">{html.escape(p)}</option>'
    for p, _ in experts
)
n_books = sum(1 for i in items if i["kind"] == "book")
n_mags = sum(1 for i in items if i["kind"] == "mag")

for tpl_name, out_dir in VARIANTS:
    with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), tpl_name), encoding="utf-8") as f:
        tpl = f.read()

    page = (tpl.replace("__CARDS__", cards_html)
               .replace("__MARQUEE__", marquee_html)
               .replace("__PERSON_OPTIONS__", persons_options)
               .replace("__TOTAL__", str(len(items)))
               .replace("__NBOOKS__", str(n_books))
               .replace("__NMAGS__", str(n_mags)))

    out = os.path.join(BASE, out_dir, "index.html")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        f.write(page)
    print(f"OK: {len(items)} изданий ({n_books} книг, {n_mags} журналов) -> {out}")
